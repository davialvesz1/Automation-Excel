import openpyxl 

#criar planilha(book)
book = openpyxl.Workbook()

#visualizar paginas existentes
print(book.sheetnames)

#criar pagina
book.create_sheet('TESTE')

#selecionar a pagina(nome da Planilha='TESTE')
test_page =  book['TESTE']
test_page.append(['Produto','Quantidade no Estoque','Preco']) 


#add dados na planilha
test_page.append(['Skate','10','RS500'])
test_page.append(['LongBoard','2','RS850'])
test_page.append(['Prancha de Surf','10','RS3200'])
test_page.append(['Patinete','7','RS3000'])
test_page.append(['Rolamento','30','RS70,00'])
test_page.append(['Rodas(skate)','48','RS179,90'])
test_page.append(['Rodas(Patinete)','22','RS249,90'])

#salvar a planilha
book.save('PlanilhaTeste.xlsx')